import requests
import pandas as pd
from datetime import datetime
import time
import re
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
# ⭐️ [제거] import io
# ⭐️ [제거] from openpyxl.drawing.image import Image

# --- ⭐️ 엑셀 스타일 및 노트 사용자 설정 ⭐️ ---
# 1. 엑셀 디자인 설정
HEADER_BG_COLOR = "#3B3EFF"
HEADER_FONT_COLOR = '#FFFFFF'
HEADER_FONT_SIZE = 12
HEADER_FONT_BOLD = True
DATA_FONT_COLOR = "#FFFFFF"
DATA_FONT_SIZE = 11
DATA_BG_COLOR_ODD = "#333333"
DATA_BG_COLOR_EVEN = '#333333'

# 2. 코인별 메모 하드코딩 설정
# '티커': '내용' 형식으로 원하는 코인과 메모를 추가하세요.
# 티커는 'USDT'를 제외한 베이스 자산명(예: 'BTC', 'ETH')을 사용합니다.
NOTE_MAP = {
    'MON': 'Pre-Market, If 10%',
    'MET': 'Pre-Market',
    '币安人生': 'bianrensheng',
    # 여기에 계속 추가 가능
}

# 3. ⭐️ [추가] 체인별 로고 하드코딩 설정 (사용자 요청)
CHAIN_LOGO_MAP = {
    'ETH': 'https://s2.coinmarketcap.com/static/img/coins/64x64/1027.png',
    'BNB': 'https://s2.coinmarketcap.com/static/img/coins/64x64/1839.png',
    'SOL': 'https://s2.coinmarketcap.com/static/img/coins/64x64/5426.png',
    'SUI': 'https://s2.coinmarketcap.com/static/img/coins/64x64/20947.png',
}

# 4. ⭐️ [추가] 하드코딩 검증 시 제외할 목록 (API가 symbol로 인식 못하는 항목)
HARDCODE_VERIFY_SKIP_LIST = ['币安人生']
# --- 설정 끝 ---

def save_to_excel(data_list, filename):
    """결과를 Excel 파일로 저장하고 사용자가 정의한 스타일을 적용합니다."""
    if not data_list:
        print("저장할 데이터가 없습니다.")
        return

    df = pd.DataFrame(data_list)
    # ⭐️ [변경] 'Chain' 열 추가로 인한 순서 조정
    column_order = ['Logo', 'Ticker', 'Name', 'Chain', 'Price', 'MarketCap', 'Listing', 'Ucid', 'Note']
    df = df.reindex(columns=column_order)

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='MarketData')

    worksheet = writer.sheets['MarketData']
    
    # ⭐️ '#' 자동 제거 로직 (이전과 동일)
    clean_header_bg = HEADER_BG_COLOR.lstrip('#')
    clean_header_font = HEADER_FONT_COLOR.lstrip('#')
    clean_data_font = DATA_FONT_COLOR.lstrip('#')
    clean_data_odd_bg = DATA_BG_COLOR_ODD.lstrip('#')
    clean_data_even_bg = DATA_BG_COLOR_EVEN.lstrip('#')

    # --- 1. 헤더 스타일 객체 생성 및 적용 ---
    header_font = Font(color=clean_header_font, size=HEADER_FONT_SIZE, bold=HEADER_FONT_BOLD)
    header_fill = PatternFill(start_color=clean_header_bg, end_color=clean_header_bg, fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    worksheet.row_dimensions[1].height = 25

    # --- 2. 열 너비 설정 ---
    worksheet.column_dimensions[get_column_letter(1)].width = 14  # Logo
    worksheet.column_dimensions[get_column_letter(2)].width = 18  # Ticker
    worksheet.column_dimensions[get_column_letter(3)].width = 24  # Name
    worksheet.column_dimensions[get_column_letter(4)].width = 14  # ⭐️ Chain (로고 표시 위해 10->14)
    worksheet.column_dimensions[get_column_letter(5)].width = 18  # Price
    worksheet.column_dimensions[get_column_letter(6)].width = 22  # MarketCap
    worksheet.column_dimensions[get_column_letter(7)].width = 12  # Listing
    worksheet.column_dimensions[get_column_letter(8)].width = 10  # Ucid
    worksheet.column_dimensions[get_column_letter(9)].width = 30  # Note
    
    # --- ⭐️ 3. [변경] 데이터 스타일 객체 사전 정의 (루프 밖에서) ---
    # (이것이 styles.xml 손상을 방지하는 핵심입니다)
    data_font = Font(color=clean_data_font, size=DATA_FONT_SIZE)
    data_fill_odd = PatternFill(start_color=clean_data_odd_bg, end_color=clean_data_odd_bg, fill_type='solid')
    data_fill_even = PatternFill(start_color=clean_data_even_bg, end_color=clean_data_even_bg, fill_type='solid')
    
    # 정렬 스타일
    base_alignment = Alignment(horizontal='left', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')
    note_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 숫자/가격 포맷
    # ⭐️ [수정] '$' 기호를 감싸는 불필요한 큰따옴표 수정
    price_format = '$#,##0.0000'
    # ⭐️ [수정] M 단위 추가 시 '$' 기호 및 소수점 누락 수정 (최종본)
    # ⭐️ [수정] T, B, M 단위의 따옴표 제거 (사용자 피드백 반영)
    marketcap_format = '[>=1000000000000]$0.00,,,, T;[>=1000000000]$0.00,,, B;[>=1000000]$0.00,, M;$#,##0'

    # --- ⭐️ 4. [변경] 데이터 행 루프 (모든 스타일을 여기서 한 번에 적용) ---
    print("엑셀 서식을 적용합니다...")
    for row_num in range(2, len(df) + 2):
        worksheet.row_dimensions[row_num].height = 70 # ⭐️ [변경] 25 -> 70 (IMAGE 함수에 맞게 높이 복원)
        
        # 교차 행 색상 결정 (2행이 1번째 데이터행(홀수))
        is_odd_row = (row_num - 1) % 2 != 0 
        fill = data_fill_odd if is_odd_row else data_fill_even

        # --- 모든 셀에 기본 스타일(font, fill, base_alignment) 적용 ---
        for col_num in range(1, 10): # 1 (A) 부터 9 (I) 까지
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = base_alignment # 기본 'left' 정렬

        # --- 특정 열 덮어쓰기 (정렬, 숫자 서식) ---
        # ⭐️ [제거] Col 1: Logo (Center) - HYPERLINK용 정렬 제거
        # worksheet.cell(row=row_num, column=1).alignment = center_alignment
        
        # Col 4: Chain (Center)
        worksheet.cell(row=row_num, column=4).alignment = center_alignment
        
        # Col 5: Price (Number Format) - 정렬은 기본(left)
        worksheet.cell(row=row_num, column=5).number_format = price_format
        
        # Col 6: MarketCap (Number Format) - 정렬은 기본(left)
        worksheet.cell(row=row_num, column=6).number_format = marketcap_format
        
        # Col 7: Listing (Center)
        worksheet.cell(row=row_num, column=7).alignment = center_alignment
        
        # Col 8: Ucid (Center)
        worksheet.cell(row=row_num, column=8).alignment = center_alignment
        
        # Col 9: Note (Wrap Text)
        worksheet.cell(row=row_num, column=9).alignment = note_alignment

    writer.close()
    print(f"\n✅ 데이터가 {filename} 파일로 저장되었습니다. (스타일 최적화 완료)")


def get_binance_cmc_data_final():
    CMC_API_KEY = "3812bfb5dd084af18752c90c94e80f9b"
    # (기존 하드코딩 목록은 그대로 유지)
    PRICE_DIFFERENCE_THRESHOLD = 0.30
    MANUAL_SUPPLY_MAP = {'1000RATS': 1_000_000_000_000, 'AVAAI': 1_000_000_000, 'MON': 10_000_000_000, 'MET': 480_000_000}
    SYMBOL_TO_ID_MAP = {
        'PUMPBTC': '36166', 'QUICK': '19966', 'ACT': '33566', 'VELODROME': '20435',
        'BROCCOLIF3B': '35753', 'BROCCOLI714': '35749', 'FXS': '6953', '币安人生': '38590',
        'MET' : '38353', 'MON' : '30495'
    }
    CMC_SKIP_LIST = []
    SPECIAL_SYMBOL_MAP = {'BEAMX': 'BEAM', 'DODOX': 'DODO', 'LUNA2': 'LUNA', 'RAYSOL': 'RAY', 'RONIN': 'RON'}
    EXCLUSION_LIST = ['BTCDOM', 'ALL']
    
    # ⭐️ [추가] 하드코딩 불일치 추적
    hardcode_discrepancies = []
    # ⭐️ [추가] 수동 시총 업데이트 가능 항목 추적
    manual_supply_updates_available = []
    
    print("데이터를 가져오는 중입니다...")

    # --- 1단계: 바이낸스 데이터 가져오기 ---
    binance_listing_dates = {}
    try:
        info_url = "https://fapi.binance.com/fapi/v1/exchangeInfo"
        info_response = requests.get(info_url); info_response.raise_for_status()
        exchange_info = info_response.json()
        active_symbols = {s['symbol'] for s in exchange_info['symbols'] if s['status'] == 'TRADING' and s['quoteAsset'] == 'USDT' and '_' not in s['symbol']}
        for s in exchange_info['symbols']:
            binance_listing_dates[s['symbol']] = s.get('onboardDate', 0)
        print(f"✅ 현재 거래 가능한 무기한 선물 페어 {len(active_symbols)}개를 확인했습니다.")
        price_url = "https://fapi.binance.com/fapi/v1/ticker/price"
        price_response = requests.get(price_url); price_response.raise_for_status()
        price_data = {item['symbol']: float(item['price']) for item in price_response.json() if item['symbol'] in active_symbols}
    except requests.exceptions.RequestException as e:
        print(f"❌ 바이낸스 API 조회 실패: {e}"); return

    # --- 2단계: CMC 조회를 위한 데이터 정규화 ---
    base_assets = {ticker.replace('USDT', '') for ticker in price_data.keys()}
    asset_to_lookup_key = {}; id_lookup_list, symbol_lookup_list = [], []
    for asset in base_assets:
        if asset in EXCLUSION_LIST or asset in CMC_SKIP_LIST: continue
        if asset in SYMBOL_TO_ID_MAP:
            cmc_id = SYMBOL_TO_ID_MAP[asset]; id_lookup_list.append(cmc_id); asset_to_lookup_key[asset] = cmc_id
        else:
            normalized = asset
            if asset.startswith(('1M','1B')): normalized = re.sub(r'^1[MB]', '', asset)
            match = re.match(r'^(10+)', asset)
            if match and len(asset) > len(match.group(0)): normalized = re.sub(r'^(10+)', '', asset)
            if asset in SPECIAL_SYMBOL_MAP: normalized = SPECIAL_SYMBOL_MAP[asset]
            symbol_lookup_list.append(normalized); asset_to_lookup_key[asset] = normalized

    # --- 3단계 (파트1): CMC 시세/시총/UCID 정보 조회 ---
    market_data_map = {} # {lookup_key: {name, market_cap, cmc_price, ucid}}
    headers = {'Accepts': 'application/json', 'X-CMC_PRO_API_KEY': CMC_API_KEY}
    
    # ID 기반 조회
    if id_lookup_list:
        print(f"ID 기반으로 {len(id_lookup_list)}개 심볼의 데이터를 조회합니다..."); time.sleep(1)
        params_str = ",".join(id_lookup_list)
        try:
            url = 'https://pro-api.coinmarketcap.com/v2/cryptocurrency/quotes/latest'
            resp = requests.get(url, headers=headers, params={'id': params_str, 'convert': 'USD'})
            resp.raise_for_status()
            for id_key, data_obj in resp.json().get('data', {}).items():
                quote = data_obj.get('quote', {}).get('USD', {})
                market_data_map[id_key] = {'name': data_obj.get('name'), 'market_cap': quote.get('market_cap'), 'cmc_price': quote.get('price'), 'ucid': id_key}
        except requests.exceptions.RequestException as e: print(f"❌ 코인마켓캡 ID API 에러: {e}")

    # Symbol 기반 조회
    symbols_to_lookup = list(set(symbol_lookup_list))
    print(f"심볼 기반으로 {len(symbols_to_lookup)}개 심볼을 조회합니다...")
    for i in range(0, len(symbols_to_lookup), 100):
        chunk = symbols_to_lookup[i:i + 100]; symbols_string = ",".join(chunk)
        try:
            url = 'https://pro-api.coinmarketcap.com/v2/cryptocurrency/quotes/latest'
            resp = requests.get(url, headers=headers, params={'symbol': symbols_string, 'convert': 'USD'})
            resp.raise_for_status()
            for symbol, data_list in resp.json().get('data', {}).items():
                if data_list:
                    info = data_list[0]; quote = info.get('quote', {}).get('USD', {})
                    ucid_from_api = str(info.get('id')) # API에서 받은 ID
                    
                    market_data_map[symbol] = {'name': info.get('name'), 'market_cap': quote.get('market_cap'), 'cmc_price': quote.get('price'), 'ucid': ucid_from_api}
            print(f" 	- {i+1} ~ {i+len(chunk)}번째 심볼 처리 완료..."); time.sleep(1)
        except requests.exceptions.RequestException as e:
            print(f"❌ 코인마켓캡 Symbol API 에러 (심볼: {symbols_string}): {e}"); continue

    # --- ⭐️ 3단계 (파트1.5): 하드코딩 맵 검증 ---
    # ⭐️ [수정] API가 'symbol' 파라미터로 인식 못하는 항목은(HARDCODE_VERIFY_SKIP_LIST) 검증에서 제외
    symbols_to_check = [symbol for symbol in SYMBOL_TO_ID_MAP.keys() if symbol not in HARDCODE_VERIFY_SKIP_LIST]
    
    if symbols_to_check:
        print(f"하드코딩된 {len(symbols_to_check)}개 심볼의 최신 정보를 검증합니다..."); time.sleep(1)
        symbols_string = ",".join(symbols_to_check)
        try:
            url = 'https://pro-api.coinmarketcap.com/v2/cryptocurrency/quotes/latest'
            resp = requests.get(url, headers=headers, params={'symbol': symbols_string, 'convert': 'USD'})
            resp.raise_for_status()
            for symbol, data_list in resp.json().get('data', {}).items():
                if data_list:
                    # 맵에 있는 심볼의 경우, API가 여러 개를 반환할 수 있음 (예: 'MET')
                    # 사용자가 하드코딩한 ID와 일치하는 항목을 찾거나, 
                    # 일치하는 것이 없으면 첫 번째 항목과 비교
                    
                    hardcoded_ucid = str(SYMBOL_TO_ID_MAP[symbol])
                    found_match = False
                    
                    for info in data_list: # API가 반환한 리스트 (e.g., 'MET' 2개)
                        api_ucid = str(info.get('id'))
                        if api_ucid == hardcoded_ucid:
                            found_match = True
                            break # ID가 일치함. 검증 통과.
                    
                    # ID가 일치하는 것을 찾지 못했다면?
                    if not found_match and data_list:
                        # API가 반환한 첫 번째 항목(가장 순위가 높은)의 ID와 비교
                        first_api_info = data_list[0]
                        first_api_ucid = str(first_api_info.get('id'))
                        
                        hardcode_discrepancies.append({
                            "symbol": symbol,
                            "hardcoded_id": hardcoded_ucid,
                            "api_id": first_api_ucid,
                            "api_name": first_api_info.get('name')
                        })
                        
        except requests.exceptions.RequestException as e:
            print(f"❌ 코인마켓캡 하드코딩 검증 API 에러 (심볼: {symbols_string}): {e}")

    # --- ⭐️ 3단계 (파트2): CMC 'Chain' 정보 조회 ---
    # (참고: 바이낸스 API는 인증된 SAPI에서만 네트워크 정보를 제공하며,
    # 이는 '네이티브 체인'이 아닌 '지원하는 모든 입출금 네트워크' 목록이라 부적합합니다.
    # 따라서 CMC의 v2/cryptocurrency/info API를 사용하는 것이 가장 정확합니다.)
    
    # 위 3-1단계에서 수집한 'ucid' 목록을 기반으로 체인 정보 조회
    ucid_list_for_chain_lookup = []
    for data in market_data_map.values():
        if data.get('ucid'):
            ucid_list_for_chain_lookup.append(str(data['ucid']))
    
    unique_ucid_list = list(set(ucid_list_for_chain_lookup))
    chain_map = {} # {ucid: chain_symbol}

    print(f"고유 ID(UCID) {len(unique_ucid_list)}개를 기반으로 'Chain' 정보를 조회합니다...")
    for i in range(0, len(unique_ucid_list), 100):
        chunk = unique_ucid_list[i:i + 100]; ids_string = ",".join(chunk)
        try:
            url = 'https://pro-api.coinmarketcap.com/v2/cryptocurrency/info'
            resp = requests.get(url, headers=headers, params={'id': ids_string})
            resp.raise_for_status()
            
            for ucid, data_obj in resp.json().get('data', {}).items():
                platform = data_obj.get('platform')
                chain_symbol = ''
                if platform:
                    chain_symbol = platform.get('symbol', '')
                else:
                    # 플랫폼 정보가 null이면, 코인 본인의 심볼을 체인으로 간주 (예: BTC, ETH)
                    chain_symbol = data_obj.get('symbol', '')
                chain_map[ucid] = chain_symbol
            print(f" 	- {i+1} ~ {i+len(chunk)}번째 ID 체인 정보 처리 완료..."); time.sleep(1)
        except requests.exceptions.RequestException as e:
            print(f"❌ 코인마켓캡 Info(Chain) API 에러 (ID: {ids_string}): {e}"); continue


    # --- 4단계: 최종 데이터 취합 ---
    final_results, problematic_tickers, price_mismatch_tickers = [], [], []
    for ticker, binance_price in price_data.items():
        base_asset = ticker.replace('USDT', '')
        if base_asset in EXCLUSION_LIST: continue
        
        lookup_key = asset_to_lookup_key.get(base_asset)
        coin_info = market_data_map.get(lookup_key)
        
        # NOTE_MAP에서 해당 코인의 메모 가져오기
        note_content = NOTE_MAP.get(base_asset, '') # 목록에 없으면 빈 문자열 반환
        
        if base_asset in CMC_SKIP_LIST:
            final_results.append({
                "Ticker": ticker, "Name": base_asset, "Price": binance_price, 
                "MarketCap": 0, "Ucid": '', "Logo": '', "Listing": '', "Chain": '',
                "Note": 'CMC_SKIP_LIST' # 스킵 목록은 별도 표기
            })
            continue
        
        if coin_info or base_asset in MANUAL_SUPPLY_MAP:
            name = coin_info.get('name', base_asset) if coin_info else base_asset
            ucid = coin_info.get('ucid', '') if coin_info else ''
            
            # ⭐️ [변경] HYPERLINK 함수 -> IMAGE 함수로 복원
            # (#NAME? 오류가 발생할 수 있습니다)
            logo_formula = f'=IMAGE("https://s2.coinmarketcap.com/static/img/coins/128x128/{ucid}.png")' if ucid else ''
            
            # ⭐️ [신규] 체인 정보 가져오기
            chain_symbol = chain_map.get(str(ucid), '') if ucid else ''
            
            # ⭐️ [추가] 체인 로고 포맷 적용 로직 (사용자 요청)
            # 맵에 심볼이 있으면 이미지로, 없으면 텍스트 심볼로
            if chain_symbol in CHAIN_LOGO_MAP:
                chain_display_value = f'=IMAGE("{CHAIN_LOGO_MAP[chain_symbol]}")'
            else:
                chain_display_value = chain_symbol
            
            listing_date = ''
            timestamp_ms = binance_listing_dates.get(ticker)
            if timestamp_ms and timestamp_ms > 0:
                listing_date = datetime.fromtimestamp(timestamp_ms / 1000).strftime('%Y-%m-%d')

            market_cap = 0
            if base_asset in MANUAL_SUPPLY_MAP:
                manual_supply = MANUAL_SUPPLY_MAP[base_asset]; multiplier = 1
                match = re.match(r'^(\d+)', base_asset)
                if match: multiplier = int(match.group(1))
                adjusted_price = binance_price / multiplier if multiplier > 0 else binance_price
                market_cap = adjusted_price * manual_supply
                
                # ⭐️ [추가] 수동 시총을 사용했으나, CMC에도 시총이 있는지 검증
                if coin_info and coin_info.get('market_cap') is not None and coin_info.get('market_cap') > 0:
                    manual_supply_updates_available.append({
                        "symbol": base_asset,
                        "manual_market_cap": market_cap,
                        "cmc_market_cap": coin_info.get('market_cap')
                    })
                    
            elif coin_info and coin_info.get('market_cap') is not None and coin_info.get('cmc_price') is not None:
                cmc_price = coin_info['cmc_price']; multiplier = 1
                if base_asset.startswith('1M'): multiplier = 1_000_000
                elif base_asset.startswith('1B'): multiplier = 1_000_000_000
                else:
                    match = re.match(r'^(10+)', base_asset)
                    if match and len(base_asset) > len(match.group(0)): multiplier = int(match.group(1))
                adjusted_binance_price = binance_price / multiplier if multiplier > 0 else binance_price
                price_diff = abs(adjusted_binance_price - cmc_price) / adjusted_binance_price if adjusted_binance_price > 0 else 1
                if price_diff > PRICE_DIFFERENCE_THRESHOLD:
                    price_mismatch_tickers.append({"ticker": ticker, "binance_price": binance_price, "cmc_price": cmc_price, "diff_percent": f"{price_diff:.1%}"})
                    continue
                else:
                    market_cap = coin_info.get('market_cap', 0)
            else:
                problematic_tickers.append(ticker); continue
                
            final_results.append({
                "Logo": logo_formula, "Ticker": ticker, "Name": name, # ⭐️ [변경] logo_link -> logo_formula
                "Chain": chain_display_value, # ⭐️ 체인 정보 (이미지 또는 텍스트)
                "Price": binance_price, "MarketCap": market_cap,
                "Listing": listing_date, # ⭐️ 'Born' -> 'Listing' 변경
                "Ucid": ucid, "Note": note_content # 하드코딩된 노트 내용 추가
            })
        else:
            problematic_tickers.append(ticker)

    # --- 5단계: 결과 저장 및 출력 ---
    final_results.sort(key=lambda x: x['MarketCap'], reverse=True)
    # ⭐️ [수정] 파일 이름 포맷 오류 수정 (%HM%S -> %H%M%S)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    
    if final_results:
        filename = f"binance_market_data_{timestamp}.xlsx"
        save_to_excel(final_results, filename)

    # ⭐️ [추가] 콘솔 출력용 시총 포맷 함수
    def format_market_cap_console(mc):
        if mc >= 1_000_000_000_000:
            return f"{mc / 1_000_000_000_000:,.2f} T"
        if mc >= 1_000_000_000:
            return f"{mc / 1_000_000_000:,.2f} B"
        if mc >= 1_000_000:
            return f"{mc / 1_000_000:,.2f} M"
        return f"{mc:,.0f}"

    # ⭐️ [변경] 출력 헤더에 'Chain' 추가, 'Born' -> 'Listing' 변경
    print(f"\n✅ 최종 {len(final_results)}개 페어 데이터 (시가총액 순, 무기한 선물만)"); print("-" * 130); print(f"{'Ticker':<18} | {'Name':<22} | {'Chain':<8} | {'Listing':<12} | {'UCID':>8} | {'Price (USD)':>18} | {'Market Cap (USD)':>22}"); print("-" * 130)
    # ⭐️ [수정] format_market_cap_console 함수 사용
    for item in final_results[:20]: print(f"{item['Ticker']:<18} | {item['Name']:<22} | {item.get('Chain', ''):<8} | {item.get('Listing', ''):<12} | {item.get('Ucid', ''):>8} | {item['Price']:>18,.4f} | {format_market_cap_console(item['MarketCap']):>22}"); print("-" * 130)
    
    if problematic_tickers:
        print("\n⚠️ 아래 티커는 코인마켓캡에서 정보를 찾을 수 없었습니다:"); print(" 	" + ", ".join(problematic_tickers))
    # ⭐️ [수정] 가격 불일치 시 CSV 파일 생성 대신 콘솔에 출력
    if price_mismatch_tickers:
        print(f"\n⚠️ {len(price_mismatch_tickers)}개의 가격 불일치 오류가 발생했습니다. (바이낸스 가격 기준)")
        print("-" * 90)
        print(f"{'Ticker':<18} | {'Binance Price':>18} | {'CMC Price':>18} | {'Difference':>15}")
        print("-" * 90)
        for item in price_mismatch_tickers:
            print(f"{item['ticker']:<18} | {item['binance_price']:>18,.4f} | {item['cmc_price']:>18,.4f} | {item['diff_percent']:>15}")
        print("-" * 90)
    
    # ⭐️ [추가] 하드코딩 맵과 API 결과 불일치 알림
    if hardcode_discrepancies:
        print("\n⚠️ 하드코딩된 'SYMBOL_TO_ID_MAP'과 CMC API 결과가 다릅니다. 업데이트를 검토하세요:")
        print("-" * 80)
        print(f"{'Symbol':<15} | {'Hardcoded ID':<15} | {'CMC API ID':<15} | {'CMC API Name'}")
        print("-" * 80)
        for item in hardcode_discrepancies:
            print(f"{item['symbol']:<15} | {item['hardcoded_id']:<15} | {item['api_id']:<15} | {item['api_name']}")
        print("-" * 80)

    # ⭐️ [추가] 수동 시총 맵 업데이트 알림
    if manual_supply_updates_available:
        print("\n⚠️ 'MANUAL_SUPPLY_MAP'의 일부 코인에 대해 CMC 공식 시총이 확인되었습니다. 수동 목록 제거를 검토하세요:")
        print("-" * 90)
        print(f"{'Symbol':<15} | {'Manual Market Cap (Calculated)':<30} | {'CMC Market Cap (Official)':<30}")
        print("-" * 90)
        for item in manual_supply_updates_available:
            print(f"{item['symbol']:<15} | {format_market_cap_console(item['manual_market_cap']):>30} | {format_market_cap_console(item['cmc_market_cap']):>30}")
        print("-" * 90)

    # ⭐️ [수정] 최종 요약 메시지
    if not problematic_tickers and not price_mismatch_tickers and not hardcode_discrepancies and not manual_supply_updates_available:
        print("\n✨ 모든 티커의 정보를 성공적으로 가져왔습니다. (모든 하드코딩 맵 일치)")
    else:
         print("\n✨ 데이터 처리를 완료했습니다. (경고 ⚠️ 항목을 확인하세요)")


if __name__ == "__main__":
    get_binance_cmc_data_final()
